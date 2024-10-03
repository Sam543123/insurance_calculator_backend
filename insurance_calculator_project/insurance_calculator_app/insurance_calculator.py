from io import BytesIO
from math import ceil, log, floor
from typing import Optional

import openpyxl
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, NamedStyle, Font, Side, Border
import datetime as dt

from openpyxl.workbook import Workbook

from insurance_calculator_app.models import LifeTable
from insurance_calculator_app.utils import format_number


class InsuranceCalculator:

    def __init__(self):
        self.v = None
        self.f = None
        self.i = None
        self.gender = None
        self.insurance_type = None
        self.insurance_premium_frequency = None
        self.lx = {}
        self.dx = {}

    def __parse_common_params(self, insurance_type: str, insurance_premium_frequency: str, gender: str,
                              insurance_premium_rate: float, insurance_loading: float):
        self.insurance_type = insurance_type
        self.insurance_premium_frequency = insurance_premium_frequency
        self.gender = gender
        self.i = insurance_premium_rate
        self.f = insurance_loading
        self.v = 1 / (1 + self.i)

    def __parse_params(self, insurance_type: str, insurance_premium_frequency: str, gender: str,
                       insurance_premium_rate: float, insurance_loading: float,
                       birth_date: dt.datetime, insurance_start_date: dt.datetime, insurance_period: Optional[int]):
        self.__parse_common_params(insurance_type, insurance_premium_frequency, gender,
                                   insurance_premium_rate, insurance_loading)
        delta = relativedelta(insurance_start_date, birth_date)
        start_age = delta.months + delta.years * 12
        processed_insurance_period = insurance_period
        if self.insurance_type == "whole life insurance":
            end_age = 1212
            processed_insurance_period = end_age - start_age
        age_lower_border = start_age // 12
        age_higher_border = (start_age + processed_insurance_period) // 12 + 1
        self.__init_life_table_metrics(age_lower_border, age_higher_border)
        return processed_insurance_period, start_age

    def __parse_tariffs_params(self, insurance_type: str, insurance_premium_frequency: str,
                               gender: str, insurance_premium_rate: float, insurance_loading: float,
                               maximum_insurance_start_age: Optional[int], minimum_insurance_start_age: Optional[int],
                               maximum_insurance_period: Optional[int]):
        self.__parse_common_params(insurance_type, insurance_premium_frequency, gender,
                                   insurance_premium_rate, insurance_loading)
        processed_maximum_insurance_period = maximum_insurance_period
        if self.insurance_type == "whole life insurance":
            end_age = 1212
            processed_maximum_insurance_period = end_age - maximum_insurance_start_age
        if self.insurance_type != "cumulative insurance":
            lower_age_border = maximum_insurance_start_age
            higher_age_border = minimum_insurance_start_age + processed_maximum_insurance_period // 12
            self.__init_life_table_metrics(lower_age_border, higher_age_border)
        return processed_maximum_insurance_period

    def __init_life_table_metrics(self, lower_age_border, higher_age_border):
        age_field = 'age'
        if self.gender == "male":
            lx_field, dx_field = 'men_survived_to_age', 'men_died_at_age'
        else:
            lx_field, dx_field = 'women_survived_to_age', 'women_died_at_age'
        filter_conditions = {f'{age_field}__gte': lower_age_border, f'{age_field}__lte': higher_age_border}
        life_table_records = LifeTable.objects.filter(**filter_conditions).values(age_field, lx_field, dx_field)
        for r in life_table_records:
            self.lx[r[age_field]] = r[lx_field]
            self.dx[r[age_field]] = r[dx_field]

    def calculate_premium(self, insurance_type: str, insurance_premium_frequency: str, gender: str,
                          insurance_premium_rate: float, insurance_loading: float, birth_date: dt.datetime,
                          insurance_start_date: dt.datetime, insurance_period: Optional[int], insurance_sum: float):
        processed_insurance_period, start_age = self.__parse_params(insurance_type, insurance_premium_frequency, gender,
                                                                    insurance_premium_rate, insurance_loading,
                                                                    birth_date,
                                                                    insurance_start_date, insurance_period)
        return self.__calculate_premium(insurance_sum, processed_insurance_period, start_age)

    def __calculate_premium(self, insurance_sum, insurance_period, start_age=None):
        sum_annuity = self.__calculate_insurance_sum_annuity(insurance_period, start_age)
        premium_annuity = self.__calculate_premium_annuity(insurance_period, start_age)
        result = (insurance_sum * sum_annuity) / (premium_annuity * (1 - self.f))
        return result

    def calculate_insurance_sum(self, insurance_type: str, insurance_premium_frequency: str, gender: str,
                                insurance_premium_rate: float, insurance_loading: float, birth_date: dt.datetime,
                                insurance_start_date: dt.datetime, insurance_period: Optional[int],
                                insurance_premium: float):
        processed_insurance_period, start_age = self.__parse_params(insurance_type, insurance_premium_frequency, gender,
                                                                    insurance_premium_rate, insurance_loading,
                                                                    birth_date, insurance_start_date, insurance_period)
        return self.__calculate_insurance_sum(insurance_premium, processed_insurance_period, start_age)

    def __calculate_insurance_sum(self, insurance_premium, insurance_period, start_age=None):
        sum_annuity = self.__calculate_insurance_sum_annuity(insurance_period, start_age)
        premium_annuity = self.__calculate_premium_annuity(insurance_period, start_age)
        result = (insurance_premium * premium_annuity * (1 - self.f)) / sum_annuity
        return result

    def calculate_reserve(self, insurance_type: str, insurance_premium_frequency: str, gender: str,
                          insurance_premium_rate: float, insurance_loading: float, birth_date: dt.datetime,
                          insurance_start_date: dt.datetime, insurance_period: Optional[int],
                          insurance_sum: Optional[float], insurance_premium: Optional[float],
                          reserve_calculation_period: int):
        processed_insurance_period, start_age = self.__parse_params(insurance_type, insurance_premium_frequency, gender,
                                                                    insurance_premium_rate, insurance_loading,
                                                                    birth_date, insurance_start_date, insurance_period)
        return self.__calculate_reserve(insurance_sum, insurance_premium, processed_insurance_period,
                                        reserve_calculation_period, start_age)

    def __calculate_reserve(self, insurance_sum, insurance_premium, insurance_period, reserve_period, start_age):
        if self.insurance_type != "cumulative insurance":
            l_t = self.__fractional_lx(start_age + reserve_period)
        if insurance_premium is None:
            insurance_premium = self.__calculate_premium(insurance_sum, insurance_period, start_age)
        else:
            insurance_sum = self.__calculate_insurance_sum(insurance_premium, insurance_period, start_age)

        if self.insurance_type in ("pure endowment", "cumulative insurance"):
            premium_annuity = self.__calculate_premium_annuity(insurance_period=reserve_period, start_age=start_age)
            if self.insurance_type == "pure endowment":
                reserve = insurance_premium * (1 - self.f) * premium_annuity / (l_t * self.v ** (reserve_period / 12))
            else:
                reserve = insurance_premium * (1 - self.f) * premium_annuity / (self.v ** (reserve_period / 12))
        else:
            premium_annuity = self.__calculate_premium_annuity(insurance_period=insurance_period,
                                                               start_age=start_age,
                                                               skip_period=reserve_period)
            sum_annuity = self.__calculate_insurance_sum_annuity(insurance_period=insurance_period,
                                                                 start_age=start_age,
                                                                 skip_period=reserve_period)
            reserve = (insurance_sum * sum_annuity - insurance_premium * (1 - self.f) * premium_annuity) / l_t
        return reserve

    def __calculate_premium_annuity(self, insurance_period, start_age=None, skip_period=0):
        annuity_cost = 0
        if self.insurance_premium_frequency == "simultaneously" and skip_period == 0:
            if self.insurance_type != "cumulative insurance":
                annuity_cost = self.__fractional_lx(start_age)
            else:
                annuity_cost = 1
        elif self.insurance_premium_frequency == "annually":
            for j, year in enumerate(range(ceil(skip_period / 12), ceil(insurance_period / 12))):
                discount_factor = self.v ** j
                if self.insurance_type != "cumulative insurance":
                    l_fraq = self.__fractional_lx(start_age + 12 * year)
                    annuity_cost += discount_factor * l_fraq
                else:
                    annuity_cost += discount_factor
        elif self.insurance_premium_frequency == "monthly":
            for j, month in enumerate(range(skip_period, insurance_period)):
                discount_factor = self.v ** (j / 12)
                if self.insurance_type != "cumulative insurance":
                    l_fraq = self.__fractional_lx(start_age + month)
                    annuity_cost += discount_factor * l_fraq
                else:
                    annuity_cost += discount_factor

        return annuity_cost

    def __calculate_insurance_sum_annuity(self, insurance_period, start_age=None, skip_period=0):
        annuity_cost = 0
        if self.insurance_type == "cumulative insurance":
            annuity_cost = self.v ** ((insurance_period - skip_period) / 12)
        elif self.insurance_type == "pure endowment":
            l_end = self.__fractional_lx(start_age + insurance_period)
            annuity_cost = self.v ** ((insurance_period - skip_period) / 12) * l_end
        else:
            for year in range((insurance_period - skip_period) // 12):
                fractional_dx = self.__fractional_dx(start_age + skip_period + 12 * year)
                annuity_cost += self.v ** (year + 1) * fractional_dx
            if self.i != 0:
                annuity_cost *= self.i / log(1 + self.i)
            if (insurance_period - skip_period) % 12 != 0:
                d_end = self.__fractional_dx(start_age + skip_period + 12 * ((insurance_period - skip_period) // 12))
                if self.i != 0:
                    annuity_cost += self.v ** ((insurance_period - skip_period) // 12 + 1) * d_end * (
                            (1 + self.i) - (1 + self.i) ** (1 - ((insurance_period - skip_period) % 12) / 12)) / log(
                        1 + self.i)
                else:
                    annuity_cost += self.v ** ((insurance_period - skip_period) // 12 + 1) * d_end * (
                            (insurance_period - skip_period) % 12) / 12

        return annuity_cost

    def __fractional_lx(self, age):
        return self.lx[age // 12] - self.dx[age // 12] * (age % 12) / 12

    def __fractional_dx(self, age):
        return self.__fractional_lx(age) - self.__fractional_lx(age + 12)

    def calculate_tariffs(self, insurance_type: str, insurance_premium_frequency: str, gender: str,
                          insurance_premium_rate: float, insurance_loading: float,
                          minimum_insurance_start_age: Optional[int], maximum_insurance_start_age: Optional[int],
                          maximum_insurance_period: int):
        processed_maximum_insurance_period = self.__parse_tariffs_params(insurance_type, insurance_premium_frequency,
                                                                         gender,
                                                                         insurance_premium_rate, insurance_loading,
                                                                         minimum_insurance_start_age,
                                                                         maximum_insurance_start_age,
                                                                         maximum_insurance_period)

        return self.__calculate_tariffs(minimum_insurance_start_age, maximum_insurance_start_age,
                                        processed_maximum_insurance_period)

    def __calculate_tariffs(self, minimum_start_age: Optional[int], maximum_start_age: Optional[int],
                            maximum_period: Optional[int]):
        if self.insurance_type != 'whole life insurance':
            tariffs_table_column_count = maximum_period // 12
        else:
            tariffs_table_column_count = 1
        tariffs_insurance_sum = 100
        tariffs_table = []
        if self.insurance_type != 'cumulative insurance':
            for x in range(minimum_start_age, maximum_start_age + 1):
                start_age = 12 * x
                tariffs_table.append([])
                for n in range(1, tariffs_table_column_count + 1):
                    if self.insurance_type == 'whole life insurance':
                        insurance_period = 1212 - 12 * x
                        tariff = format_number(self.__calculate_premium(tariffs_insurance_sum, insurance_period, start_age))
                    elif self.insurance_type == 'whole life insurance' or x + n <= 101:
                        insurance_period = 12 * n
                        tariff = format_number(self.__calculate_premium(tariffs_insurance_sum, insurance_period, start_age))
                    else:
                        tariff = '-'
                    tariffs_table[-1].append(tariff)

        else:
            for n in range(maximum_period // 12 + 1):
                tariffs_table.append([])
                for m in range(12):
                    insurance_period = 12 * n + m
                    if m + n != 0 and 12 * n + m <= maximum_period:
                        tariff = format_number(self.__calculate_premium(tariffs_insurance_sum, insurance_period))
                    else:
                        tariff = '-'
                    tariffs_table[-1].append(tariff)

        file = self.__create_tariffs_table(tariffs_table, minimum_start_age)
        return file

    def __create_tariffs_table(self, tariffs_table, start_age):
        row_number = len(tariffs_table)
        column_number = len(tariffs_table[0])
        tariffs_page = Workbook()
        work_sheet = tariffs_page.active
        work_sheet.title = 'Tariffs'

        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, top=border_side, right=border_side, bottom=border_side)
        header_font = Font(name='Times New Roman', size=14)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        table_text_style = NamedStyle(name='tableText', font=Font(name='Times New Roman', size=12),
                                      alignment=Alignment(horizontal='center', vertical='center'),
                                      border=border)
        table_header_style = NamedStyle(name='tableHeader', font=header_font,
                                        alignment=center_alignment,
                                        border=border)

        header_style = NamedStyle(name='header', font=header_font, alignment=center_alignment)

        tariffs_page.add_named_style(header_style)
        tariffs_page.add_named_style(table_header_style)
        tariffs_page.add_named_style(table_text_style)

        for i in range(1, 5):
            work_sheet[f'A{i}'].style = 'header'
            work_sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=column_number + 1)

        work_sheet['A5'].style = 'tableHeader'
        work_sheet['B5'].style = 'tableHeader'

        work_sheet['A1'].font = Font(name='Times New Roman', size=14, bold=True)
        work_sheet[
            'A1'] = f'Tariffs table in % with insurance premium rate {(100 * self.i):.2f}% and loading {(100 * self.f):.2f}%'
        work_sheet['A2'] = f'Insurance type: {self.insurance_type}'
        work_sheet['A3'] = f'Payment frequency: {self.insurance_premium_frequency}'

        work_sheet.row_dimensions[1].height = 50
        work_sheet.row_dimensions[2].height = 40
        work_sheet.row_dimensions[3].height = 40
        work_sheet.row_dimensions[4].height = 40

        skipped_rows = 6
        if self.insurance_type != 'cumulative insurance':
            work_sheet.column_dimensions['A'].width = 20
            work_sheet['A4'] = f'Gender of the insured person: {self.gender}'
            work_sheet['A5'] = 'Age of the insured person'
            if self.insurance_type != 'whole life insurance':
                work_sheet.row_dimensions[5].height = 30
                work_sheet.row_dimensions[6].height = 20
                work_sheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
                work_sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=column_number + 1)
                work_sheet['B5'] = 'Insurance period (years)'
                for j in range(1, column_number + 1):
                    cell = work_sheet.cell(row=6, column=j + 1)
                    cell.style = 'tableText'
                    cell.value = j
            else:
                work_sheet.row_dimensions[5].height = 50
                work_sheet.column_dimensions['B'].width = 20
                work_sheet['B5'] = 'Tariff'
                skipped_rows = 5

            start_first_column_value = start_age

        else:
            work_sheet.row_dimensions[5].height = 30
            work_sheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
            work_sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=column_number + 1)
            work_sheet['A4'] = 'Insurance period (years, months)'
            work_sheet['A5'] = 'Year'
            work_sheet['B5'] = 'Month'

            for j in range(2, column_number + 2):
                cell = work_sheet.cell(row=6, column=j)
                cell.style = 'tableText'
                cell.value = j - 2
            start_first_column_value = 0

        for i in range(1, row_number + 1):
            cell = work_sheet.cell(row=i + skipped_rows, column=1)
            cell.style = 'tableText'
            cell.value = start_first_column_value + (i - 1)
            for j in range(2, column_number + 2):
                cell = work_sheet.cell(row=i + skipped_rows, column=j)
                cell.style = 'tableText'
                cell.value = tariffs_table[i - 1][j - 2]

        # Saving Excel file in memory
        output = BytesIO()
        tariffs_page.save(output)
        output.seek(0)
        return output
